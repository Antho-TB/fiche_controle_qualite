import pandas as pd
import openpyxl
from datetime import datetime
import os
import sys

# Configuration
CSV_PATH = "article.csv"
EXCEL_TEMPLATE = "FOR-ACH-30-2 Fiche d'inspection produit-Contrôle réception.xlsx"
OUTPUT_DIR = "fiches_generees"

def load_articles():
    """Charge le CSV des articles avec gestion du délimiteur ;"""
    try:
        # On lit les 3 colonnes : Référence, EAN (Gencod), Désignation
        df = pd.read_csv(CSV_PATH, sep=';', encoding='utf-8', dtype=str, header=None)
        df.columns = ['ref', 'ean', 'designation']
        # On nettoie les espaces
        df['ref'] = df['ref'].str.strip()
        df['ean'] = df['ean'].str.strip()
        df['designation'] = df['designation'].str.strip()
        return df
    except Exception as e:
        print(f"Erreur lors du chargement de {CSV_PATH}: {e}")
        return None

def find_article(df, code):
    """Recherche un article par EAN ou Référence"""
    code = str(code).strip()
    # Recherche par EAN
    res = df[df['ean'] == code]
    if res.empty:
        # Recherche par Référence
        res = df[df['ref'] == code]
    
    if not res.empty:
        return res.iloc[0].to_dict()
    return None

def update_excel(article_info, quantity=1):
    """Met à jour le fichier Excel avec les infos de l'article"""
    if not os.path.exists(OUTPUT_DIR):
        os.makedirs(OUTPUT_DIR)
    
    # Nom du fichier de sortie basé sur la date du jour si on veut un nouveau fichier par session
    date_str = datetime.now().strftime("%Y-%m-%d")
    output_path = os.path.join(OUTPUT_DIR, f"Fiche_Controle_{date_str}.xlsx")
    
    # Si le fichier n'existe pas, on part du template
    if not os.path.exists(output_path):
        wb = openpyxl.load_workbook(EXCEL_TEMPLATE)
    else:
        wb = openpyxl.load_workbook(output_path)
    
    ws = wb.active # On suppose que c'est la première feuille

    # Mapping des cellules selon l'analyse précédente (Feuil1)
    # D'après l'inspection :
    # Cell(9, 4) [D9] semble être "Référence Articles :" - On met la valeur à côté ou dessous ?
    # Habituellement dans ces fiches :
    # - Date de réception : B10 ou similaire
    # - Référence : D9
    # - Désignation : D10
    
    # On va remplir les champs identifiés
    ws['D9'] = article_info['ref']
    ws['D10'] = article_info['designation']
    ws['D11'] = datetime.now().strftime("%d/%m/%Y") # Date de réception
    
    # On peut aussi ajouter le code EAN si besoin
    # ws['D12'] = article_info['ean']

    wb.save(output_path)
    return output_path

def main():
    print("=== Outil Douchette Qualité ===")
    print(f"Chargement de la base article ({CSV_PATH})...")
    df_articles = load_articles()
    
    if df_articles is None:
        return

    print("Prêt ! Scannez un code-barre ou tapez une référence (tapez 'exit' pour quitter).")
    
    while True:
        try:
            code = input("\nScan / Code article : ").strip()
            if code.lower() in ['exit', 'quit', 'q']:
                break
            
            if not code:
                continue
                
            article = find_article(df_articles, code)
            
            if article:
                print(f"Trouvé : {article['designation']} (Réf: {article['ref']})")
                path = update_excel(article)
                print(f"Fiche Excel mise à jour : {path}")
            else:
                print(f"[-] Code '{code}' inconnu dans la base article.")
        except KeyboardInterrupt:
            break
        except Exception as e:
            print(f"Erreur : {e}")

if __name__ == "__main__":
    main()
