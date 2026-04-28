"""
[ARCHITECTURE] Génération Documentaire (fiche_de_controle)

Rôle global :
Ce module gère l'injection des données consolidées (issues de Sylob, du CSV et de l'extraction PDF)
directement dans un template Excel d'inspection qualité. C'est l'interface de sortie vers les équipes
opérationnelles (Qualité / Réception).

Stratégie métier (Template injection) :
Il est crucial d'utiliser la librairie `openpyxl` plutôt que `pandas` pour l'export. Pandas
écraserait les macros, la mise en forme (couleurs, polices, tailles de cellules) et les 
formules pré-existantes dans le fichier modèle (FOR-ACH-30-2). En utilisant openpyxl, on 
édite chirurgicalement des cellules spécifiques (ex: B4, G5) tout en conservant l'intégrité
du document qualité certifié.
"""

import openpyxl
from datetime import datetime
import os
import sys
import logging

def get_base_path() -> str:
    """
    Retourne le chemin d'exécution absolu.
    
    Stratégie :
    Prend en compte l'exécution depuis un exécutable compilé (PyInstaller)
    pour garantir que le dossier racine est toujours correctement résolu, 
    empêchant les erreurs "File Not Found" sur les postes des opérateurs.
    """
    if getattr(sys, 'frozen', False):
        return os.path.dirname(sys.executable)
    return os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))


class ExcelHandler:
    """
    Classe utilitaire gérant la manipulation du fichier Excel de contrôle réception.
    """
    
    def __init__(self, template_path: str = None):
        base_path = get_base_path()
        if template_path is None:
            self.template_path = os.path.join(base_path, "0_Modele_Et_Donnees", "FOR-ACH-30-2 Fiche d'inspection produit-Contrôle réception.xlsx")
        else:
            self.template_path = template_path
            
        self.output_dir = os.path.join(base_path, "2_Fiches_Creees")
        
        # S'assurer que le dossier de sortie existe
        if not os.path.exists(self.output_dir):
            os.makedirs(self.output_dir)

    def generer_fiche(self, article_info: dict) -> str:
        """
        Remplit une nouvelle fiche basée sur le template Excel avec les données de l'article.
        
        Stratégie :
        - Chargement du fichier original.
        - Injection par coordonnées exactes (ex: B5 pour la référence).
        - Nettoyage préventif (colonne H) pour éviter des reliquats de données si 
          le template a été mal sauvegardé précédemment.
        - Sauvegarde sous un nom horodaté (timestamp) pour éviter tout écrasement accidentel.
        
        Args:
            article_info (dict): Les métadonnées consolidées de l'article.
            
        Returns:
            str: Le chemin d'accès absolu au fichier Excel fraîchement généré (ou None si échec).
        """
        if not os.path.exists(self.template_path):
            logging.error(f"[ERREUR] Template Excel introuvable : {self.template_path}")
            return None

        # Stratégie de nommage (Traçabilité)
        # Format : Fiche_[Ref]_[Lot]_[Timestamp].xlsx
        now = datetime.now()
        timestamp = now.strftime("%Y%m%d_%H%M%S")
        
        # Le lot peut contenir des caractères problématiques, on s'assure qu'il est propre 
        lot = str(article_info.get('lot', '')).replace("/", "-").replace("\\", "-")
        lot_suffix = f"_{lot}" if lot else ""
        
        nom_sortie = f"Fiche_{article_info['ref']}{lot_suffix}_{timestamp}.xlsx"
        chemin_sortie = os.path.join(self.output_dir, nom_sortie)

        try:
            # openpyxl conserve la mise en forme du template
            wb = openpyxl.load_workbook(self.template_path)
            ws = wb.active
            
            # --- Injection des données métier ---
            from openpyxl.styles import Font
            red_font = Font(color="FF0000")
            
            # Fonction utilitaire pour injecter en rouge
            def set_red_value(cell_coord, value):
                ws[cell_coord] = value
                ws[cell_coord].font = red_font
                
            # 1. Dates (Format FR standard)
            set_red_value('B4', now.strftime("%d/%m/%Y"))
            set_red_value('F4', now.strftime("%d/%m/%Y"))
            
            # 2. Identification Produit
            set_red_value('B5', article_info['ref'])
            set_red_value('B6', article_info['designation'])

            # 3. Traçabilité
            set_red_value('G5', article_info.get('po', ''))
            set_red_value('G6', lot)
            
            # 4. Nettoyage préventif de la colonne H (Commentaires du template vierge)
            for row in range(5, 51):
                ws[f'H{row}'] = None
                
            # 5. Injection conditionnelle Fournisseur et EANs supplémentaires
            fournisseur = article_info.get('fournisseur', '')
            if fournisseur:
                set_red_value('C9', fournisseur)
                ws.merge_cells('C9:G9')
                
            ean_spcb = str(article_info.get('ean_spcb', '')).replace('nan', '').strip()
            ean_pcb = str(article_info.get('ean_pcb', '')).replace('nan', '').strip()
            ho = str(article_info.get('ho', '')).replace('nan', '').strip()
            
            if ean_pcb:
                set_red_value('F12', ean_pcb)
            if ean_spcb:
                set_red_value('F14', ean_spcb)
            if ho:
                set_red_value('B35', ho)
            
            # Sérialisation
            wb.save(chemin_sortie)
            logging.info(f"[SUCCÈS] Fiche générée : {chemin_sortie}")
            return chemin_sortie

        except Exception as e:
            # Fallback (si le fichier est ouvert par un autre processus par exemple)
            logging.error(f"[ERREUR] Échec de manipulation du fichier Excel : {e}", exc_info=True)
            return None

if __name__ == "__main__":
    # Test unitaire rapide
    h = ExcelHandler()
    dump_article = {'ref': 'TEST-123', 'designation': 'Article de Test', 'ean': '000000'}
    h.generer_fiche(dump_article)
