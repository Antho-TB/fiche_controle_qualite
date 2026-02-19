import pandas as pd
import openpyxl

def detailed_excel_inspect(path):
    # Read with pandas to see content
    df = pd.read_excel(path, sheet_name=0, header=None)
    print("Full Content (first 30 rows):")
    print(df.iloc[:30, :10].to_string())

    # Read with openpyxl to see merged cells or specific styles if needed
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    print("\nSpecific Cell Values:")
    for row in range(1, 15):
        for col in range(1, 10):
            val = ws.cell(row=row, column=col).value
            if val:
                print(f"Cell({row}, {col}) [{openpyxl.utils.get_column_letter(col)}{row}]: {val}")

if __name__ == "__main__":
    detailed_excel_inspect("FOR-ACH-30-2 Fiche d'inspection produit-Contrôle réception.xlsx")
